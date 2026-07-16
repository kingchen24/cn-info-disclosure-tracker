"""把示例 Excel 渲染成 PNG（用于 README 截图）。
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

XLSX = ROOT / "output" / "sample.xlsx"
OUT_DIR = ROOT / "docs" / "screenshots"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def find_font(candidates, size=14):
    for fp in candidates:
        if Path(fp).exists():
            return ImageFont.truetype(fp, size)
    return ImageFont.load_default()


FONT_CANDIDATES_WINDOWS = [
    r"C:\Windows\Fonts\msyh.ttc",
    r"C:\Windows\Fonts\msyh.ttf",
    r"C:\Windows\Fonts\simhei.ttf",
    r"C:\Windows\Fonts\simsun.ttc",
]
FONT = find_font(FONT_CANDIDATES_WINDOWS, 13)
FONT_BOLD = find_font(FONT_CANDIDATES_WINDOWS, 14)
FONT_GROUP = find_font(FONT_CANDIDATES_WINDOWS, 15)

COL_WIDTHS = [44, 70, 70, 200, 70, 200, 100, 280, 70, 80, 70, 90, 200, 110, 260]
HEAD_FILL = (48, 84, 150)
HEAD_FG = (255, 255, 255)
GROUP_FILL = (210, 222, 245)
ROW_ALT = (240, 244, 252)
BORDER = (200, 200, 200)


def text_w(draw, s, font):
    return draw.textbbox((0, 0), s or "", font=font)[2]


def text_lines(draw, s, font, max_w):
    lines, cur = [], ""
    for ch in s:
        if text_w(draw, cur + ch, font) > max_w:
            lines.append(cur); cur = ch
        else:
            cur += ch
    if cur: lines.append(cur)
    return lines


def render_table(rows, widths, out_path, title=""):
    """rows: [(kind, [col, col, ...])]
       kind: 'group' / 'data'
    """
    cell_h = 30
    pad_x = 8
    pad_y = 5
    headers = ["项目", "数量"]

    tmp = Image.new("RGB", (10, 10))
    d = ImageDraw.Draw(tmp)
    col_w = []
    for i, w in enumerate(widths):
        max_text = headers[i] if i < len(headers) else ""
        for kind, r in rows:
            if i < len(r):
                v = str(r[i])[:50]
                if text_w(d, v, FONT) > text_w(d, max_text, FONT_BOLD):
                    max_text = v
        col_w.append(max(w, text_w(d, max_text, FONT_BOLD) + 16))

    table_w = sum(col_w)
    table_h = cell_h * (1 + len(rows)) + (cell_h + 6 if title else 0)
    img = Image.new("RGB", (table_w + 40, table_h + 40), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    x0, y0 = 20, 20

    if title:
        draw.text((x0, y0), title, fill=(40, 40, 40), font=FONT_BOLD)
        y0 += cell_h + 6

    cur_x = x0
    for i, h in enumerate(headers):
        draw.rectangle([cur_x, y0, cur_x + col_w[i], y0 + cell_h], fill=HEAD_FILL, outline=BORDER)
        draw.text((cur_x + pad_x, y0 + pad_y), h, fill=HEAD_FG, font=FONT_BOLD)
        cur_x += col_w[i]
    cur_y = y0 + cell_h

    for ri, (kind, r) in enumerate(rows):
        row_height = cell_h
        if kind != "group":
            for i in range(len(headers)):
                lines = text_lines(draw, str(r[i]) if i < len(r) else "", FONT, col_w[i] - 16)
                row_height = max(row_height, len(lines) * (FONT.size + 4) + pad_y * 2)

        cur_x = x0
        if kind == "group":
            for i in range(len(headers)):
                draw.rectangle([cur_x, cur_y, cur_x + col_w[i], cur_y + row_height],
                               fill=GROUP_FILL, outline=BORDER)
                if i == 0:
                    draw.text((cur_x + pad_x, cur_y + pad_y), str(r[i]),
                              fill=(40, 60, 120), font=FONT_GROUP)
                cur_x += col_w[i]
        else:
            bg = ROW_ALT if ri % 2 == 0 else (255, 255, 255)
            for i in range(len(headers)):
                val = str(r[i]) if i < len(r) else ""
                draw.rectangle([cur_x, cur_y, cur_x + col_w[i], cur_y + row_height],
                               fill=bg, outline=BORDER)
                lines = text_lines(draw, val, FONT, col_w[i] - 16)
                for li, ln in enumerate(lines[:3]):
                    draw.text((cur_x + pad_x, cur_y + pad_y + li * (FONT.size + 4)),
                              ln, fill=(40, 40, 40), font=FONT)
                cur_x += col_w[i]
        cur_y += row_height

    img.save(out_path, "PNG", optimize=True)
    print(f"[OK] {out_path}  size={img.size}")


def render_main_sheet(rows, headers, widths, out_path, title=""):
    cell_h = 30
    pad_x = 8
    pad_y = 5

    tmp = Image.new("RGB", (10, 10))
    d = ImageDraw.Draw(tmp)
    col_w = []
    for i, w in enumerate(widths):
        h_text = headers[i] if i < len(headers) else ""
        max_text = h_text
        for r in rows:
            if i < len(r):
                v = str(r[i])[:60]
                if text_w(d, v, FONT) > text_w(d, max_text, FONT_BOLD):
                    max_text = v
        col_w.append(max(w, text_w(d, max_text, FONT_BOLD) + 16))

    table_w = sum(col_w)
    img = Image.new("RGB", (table_w + 40, 99999), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    x0, y0 = 20, 20
    if title:
        draw.text((x0, y0), title, fill=(40, 40, 40), font=FONT_BOLD)
        y0 += cell_h + 6

    cur_x = x0
    for i, h in enumerate(headers):
        draw.rectangle([cur_x, y0, cur_x + col_w[i], y0 + cell_h], fill=HEAD_FILL, outline=BORDER)
        draw.text((cur_x + pad_x, y0 + pad_y), h, fill=HEAD_FG, font=FONT_BOLD)
        cur_x += col_w[i]
    cur_y = y0 + cell_h

    for ri, r in enumerate(rows):
        row_height = cell_h
        for i in range(len(headers)):
            val = str(r[i]) if i < len(r) else ""
            lines = text_lines(draw, val, FONT, col_w[i] - 16)
            row_height = max(row_height, len(lines) * (FONT.size + 4) + pad_y * 2)
        cur_x = x0
        bg = ROW_ALT if ri % 2 == 0 else (255, 255, 255)
        for i in range(len(headers)):
            val = str(r[i]) if i < len(r) else ""
            draw.rectangle([cur_x, cur_y, cur_x + col_w[i], cur_y + row_height],
                           fill=bg, outline=BORDER)
            lines = text_lines(draw, val, FONT, col_w[i] - 16)
            for li, ln in enumerate(lines[:3]):
                draw.text((cur_x + pad_x, cur_y + pad_y + li * (FONT.size + 4)),
                          ln, fill=(40, 40, 40), font=FONT)
            cur_x += col_w[i]
        cur_y += row_height

    img = img.crop((0, 0, img.size[0], cur_y + 20))
    img.save(out_path, "PNG", optimize=True)
    print(f"[OK] {out_path}  size={img.size}")


def main():
    wb = openpyxl.load_workbook(XLSX)

    # ---- 主 sheet ----
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=13):
        rows.append(["" if c.value is None else str(c.value) for c in row])
    render_main_sheet(rows, headers, COL_WIDTHS,
                      OUT_DIR / "main-sheet.png",
                      title="主 sheet：半年报业绩预告明细（主板）")

    # ---- 汇总 sheet ----
    ws2 = wb["汇总"]
    rows2 = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        a = row[0]
        b = row[1] if len(row) > 1 else None
        if a is None: continue
        a = str(a)
        b_str = "" if b is None else str(b)
        if a.endswith("股票数"):
            rows2.append(("group", [a, b_str]))
        elif "股票数（去重）" in a:
            rows2.append(("data", [a, b_str]))
        else:
            rows2.append(("data", [a, b_str]))
    render_table(rows2, [380, 100],
                 OUT_DIR / "summary-sheet.png",
                 title="汇总 sheet：行业 / 板块分布")


if __name__ == "__main__":
    main()
