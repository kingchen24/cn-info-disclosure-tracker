"""把合并后的 DataFrame 输出成漂亮的 Excel。

- 主 sheet：明细（带超链接公告）
- 汇总 sheet：行业分布 / 板块分布
- 表头深蓝底白字、冻结首行、自动筛选
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "序号", "股票代码", "股票简称", "公司全称",
    "所属市场", "所属行业", "上市日期",
    "主营业务简介", "市盈率-静", "市盈率-TTM", "市净率", "总市值(亿元)",
    "公告标题", "公告时间", "公告链接",
]
WIDTHS = [6, 10, 10, 28, 12, 24, 12, 46, 11, 12, 9, 14, 30, 18, 46]


def _fmt_num(x, kind: str) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or x in ("", "nan"):
        return ""
    try:
        if kind == "pe":
            return f"{float(x):.2f}"
        if kind == "mv":
            return f"{float(x):,.2f}"
    except (TypeError, ValueError):
        return str(x)
    return str(x)


def export(df: pd.DataFrame, summary_dfs: dict, xlsx_path: Path, cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{cfg.get('period','')}{cfg.get('disclosure','')}-{cfg.get('board_filter','')}"

    # 表头
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 数据行
    for _, row in df.iterrows():
        biz = row.get("主营业务简介") or ""
        if len(biz) > 80:
            biz = biz[:78] + ".."
        ws.append([
            row.get("序号"),
            row.get("股票代码"),
            row.get("股票简称"),
            row.get("公司全称"),
            row.get("所属市场"),
            row.get("所属行业"),
            row.get("上市日期"),
            biz,
            _fmt_num(row.get("市盈率-静"), "pe"),
            _fmt_num(row.get("市盈率-TTM"), "pe"),
            _fmt_num(row.get("市净率"), "pe"),
            _fmt_num(row.get("总市值(亿元)"), "mv"),
            row.get("公告标题"),
            row.get("公告时间"),
            row.get("公告链接"),
        ])
        r = ws.max_row
        url = row.get("公告链接") or ""
        cell = ws.cell(row=r, column=len(HEADERS))
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (8, 13)))
            c.border = border

    # 列宽 + 冻结 + 筛选
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36

    # 汇总 sheet
    ws2 = wb.create_sheet("汇总")
    ws2.append([f"{cfg.get('period','')}{cfg.get('disclosure','')} 汇总", ""])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.append([])
    ws2.append(["统计项", "数量"])
    for col in (1, 2):
        c = ws2.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ws2.append([f"{cfg.get('board_filter','')}股票数（去重）", len(df)])

    ws2.append([])
    ws2.append(["所属行业", "股票数"])
    for col in (1, 2):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for k, v in summary_dfs.get("industry", pd.DataFrame()).itertuples(index=False):
        ws2.append([k, int(v)])

    ws2.append([])
    ws2.append(["所属市场", "股票数"])
    for col in (1, 2):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for k, v in summary_dfs.get("market", pd.DataFrame()).itertuples(index=False):
        ws2.append([k, int(v)])

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 14
    ws2.freeze_panes = "A2"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
